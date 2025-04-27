import os
import random
import requests
from openpyxl import load_workbook
from telegram import Update, InputFile
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

# ─── ПЕРЕМЕННЫЕ ОКРУЖЕНИЯ ───────────────────────────────────────────────────────
ACCESS_TOKEN      = os.getenv("ACCESS_TOKEN")       # Page Access Token для Instagram Business API
MEDIA_ID          = os.getenv("MEDIA_ID")           # ID поста в Instagram
TELEGRAM_TOKEN    = os.getenv("TELEGRAM_TOKEN")     # Токен вашего Telegram-бота
DOWNLOAD_PASSWORD = os.getenv("DOWNLOAD_PASSWORD")  # Пароль для команды /download

EXCEL_FILE = "promo_codes_test.xlsx"
SHEET_NAME = "Лист1"

# ─── СООБЩЕНИЯ ─────────────────────────────────────────────────────────────────
START_MESSAGE = """Привет! 👋  
Отправь мне свой Instagram-никнейм (например, @yourname), я проверю комментарий под нашим постом и выдам промокод."""
ASK_PASSWORD_MESSAGE   = "🔐 Введите пароль для скачивания файла:"
WRONG_PASSWORD_MESSAGE = "🚫 Неверный пароль."
FILE_NOT_FOUND_MESSAGE = "🚫 Файл не найден."
ALREADY_GOT_MESSAGE    = "✅ Вы уже получили промокод: *{promo_code}*"
SUCCESS_TEMPLATE       = """✅ Отлично! Комментарий обнаружен.
Ваш промокод: *{promo_code}*
Используйте его до {expiry}."""
FAIL_MESSAGE           = """😕 Комментарий под постом не найден. Проверьте, пожалуйста:
1. Подписка на @aviashow.kz
2. Лайк на пост
3. Комментарий с отметкой двух друзей"""

# ─── ФУНКЦИИ РАБОТЫ С EXCEL ────────────────────────────────────────────────────
def find_user_in_sheet(username: str):
    """Ищет в колонке D (4) никнейм. Возвращает (row, promo_code) или None."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, _, expiry, used = row[0], row[1], row[2], row[3]
        if used and used.lower() == username.lower():
            wb.close()
            return (row, code)
    wb.close()
    return None

def get_available_codes():
    """Возвращает список (promo_code, row_number) для пустых Used."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    out = []
    for row in ws.iter_rows(min_row=2):
        code_cell = row[0]
        used_cell = row[3]
        if code_cell.value and (used_cell.value is None or used_cell.value == ""):
            out.append((code_cell.value, code_cell.row, ws.cell(row=code_cell.row, column=3).value))
    wb.close()
    return out

def mark_code_as_used(row: int, username: str):
    """Записывает username в колонку D на строке row."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.cell(row=row, column=4, value=username)
    wb.save(EXCEL_FILE)
    wb.close()

# ─── ФУНКЦИЯ ПРОВЕРКИ КОММЕНТАРИЯ В INSTAGRAM ──────────────────────────────────
def has_user_commented(username: str) -> bool:
    url = f"https://graph.facebook.com/v19.0/{MEDIA_ID}/comments?fields=username,text"
    params = {
        "access_token": ACCESS_TOKEN,
        "fields": "owner.username,text",
        "limit": 100,
    }
    while url:
        resp = requests.get(url, params=params).json()
        for c in resp.get("data", []):
            owner = c.get("owner", {})
            if owner.get("username", "").lower() == username.lower():
                return True
        url = resp.get("paging", {}).get("next")
    return False

# ─── ХАНДЛЕРЫ ТЕЛЕГРАМ ────────────────────────────────────────────────────────
async def download_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Обрабатывает /download — спрашивает пароль."""
    context.user_data["await_download"] = True
    await update.message.reply_text(ASK_PASSWORD_MESSAGE)

async def text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    user = update.effective_user.id

    # — если ждём пароль для скачивания
    if context.user_data.get("await_download"):
        context.user_data["await_download"] = False
        if text == DOWNLOAD_PASSWORD:
            if os.path.exists(EXCEL_FILE):
                await update.message.reply_document(InputFile(EXCEL_FILE, filename="promo_codes.xlsx"))
            else:
                await update.message.reply_text(FILE_NOT_FOUND_MESSAGE)
        else:
            await update.message.reply_text(WRONG_PASSWORD_MESSAGE)
        return

    # — иначе считаем, что это Instagram-никнейм
    username = text.lstrip("@")
    await update.message.reply_text(f"🔍 Проверяю комментарий от @{username}…")

    # 1) проверим, не получал ли уже пользователь код
    found = find_user_in_sheet(username)
    if found:
        _, promo_code = found
        await update.message.reply_text(ALREADY_GOT_MESSAGE.format(promo_code=promo_code), parse_mode="Markdown")
        return

    # 2) проверяем комментарий в Instagram
    if not has_user_commented(username):
        await update.message.reply_text(FAIL_MESSAGE)
        return

    # 3) выдаём новый код
    available = get_available_codes()
    if not available:
        await update.message.reply_text("😔 Промокоды закончились.")
        return

    promo_code, row, expiry = random.choice(available)
    mark_code_as_used(row, username)
    await update.message.reply_text(
        SUCCESS_TEMPLATE.format(promo_code=promo_code, expiry=expiry),
        parse_mode="Markdown"
    )

# ─── СТАРТ БОТА ───────────────────────────────────────────────────────────────
def main():
    app = ApplicationBuilder().token(TELEGRAM_TOKEN).build()

    # команда /download
    app.add_handler(CommandHandler("download", download_command))
    # всё остальное — текстовые сообщения
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, text_handler))

    app.run_polling()

if __name__ == "__main__":
    main()
